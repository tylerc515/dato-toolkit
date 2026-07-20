"""Tests for TEAM xlsx parser. Tests run against real files in examples/team/."""
from __future__ import annotations

import dataclasses

import openpyxl
import pytest

FLOOR = "examples/team/FLOOR.xlsx"
FLOOR_MLO = "examples/team/FLOOR MLO.xlsx"


def test_no_metadata_fields_on_result():
    from app.converters.team_parser import parse_team_file

    result = parse_team_file(FLOOR)
    field_names = {f.name for f in dataclasses.fields(result)}
    metadata_fields = {
        "company_name",
        "mill_location",
        "boiler_name",
        "inspection_date",
        "boiler_section",
        "nde_laboratory",
    }
    assert field_names.isdisjoint(metadata_fields)


def test_left_to_right_direction():
    from app.converters.team_parser import parse_team_file

    result = parse_team_file(FLOOR)
    assert result.tube_numbers[0] == 1
    assert result.numbering_direction == "Left-to-Right"


def test_right_to_left_direction_synthetic(tmp_path):
    from app.converters.team_parser import parse_team_file

    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1: tube numbers starting at column C, descending (Right-to-Left)
    for i, tube in enumerate((5, 4, 3, 2, 1)):
        ws.cell(row=1, column=3 + i, value=tube)
    # One 3-row elevation block with real readings
    ws.cell(row=2, column=1, value="Test Elevation")
    ws.cell(row=2, column=2, value="L")
    ws.cell(row=3, column=2, value="C")
    ws.cell(row=4, column=2, value="R")
    for i in range(5):
        ws.cell(row=2, column=3 + i, value=100 + i)
        ws.cell(row=3, column=3 + i, value=200 + i)
        ws.cell(row=4, column=3 + i, value=300 + i)

    path = tmp_path / "synthetic_rtl.xlsx"
    wb.save(path)

    result = parse_team_file(path)
    assert result.tube_numbers[0] == 5
    assert result.numbering_direction == "Right-to-Left"


def test_parser_returns_all_blocks_with_has_data():
    from app.converters.team_parser import parse_team_file

    result = parse_team_file(FLOOR_MLO)
    # FLOOR MLO has 23 total template block positions; only 2 are measured.
    assert len(result.elevations) == 23

    measured = [e for e in result.elevations if e.has_data]
    blank = [e for e in result.elevations if not e.has_data]
    assert len(measured) == 2
    assert len(blank) == 21

    measured_labels = {e.label for e in measured}
    assert measured_labels == {
        'Floor to Wall Bend Weld Line Rear Wall +1"',
        'Floor to Wall Bend Weld Line Frnt Wall +1"',
    }


def test_reading_zero_padding():
    from app.converters.team_parser import convert_team_reading

    assert convert_team_reading(58) == "058"
    assert convert_team_reading(214) == "214"
    assert convert_team_reading(None) == ""


def test_flag_passthrough():
    from app.converters.team_parser import convert_team_reading

    assert convert_team_reading("*") == "*"
    assert convert_team_reading("<") == "<"


def test_suffix_letter_preserved():
    from app.converters.team_parser import convert_team_reading

    assert convert_team_reading("14V") == "014V"


def test_out_of_range_value_passes_through_unmodified():
    from app.converters.team_parser import convert_team_reading

    assert convert_team_reading(2791) == "2791"


def test_comment_codes_found_collected():
    from app.converters.team_parser import parse_team_file

    result = parse_team_file(FLOOR)
    assert result.comment_codes_found == {"<"}


def test_parse_invalid_file_raises():
    from app.converters.team_parser import TEAMParseError, parse_team_file

    with pytest.raises(TEAMParseError):
        parse_team_file("nonexistent_file_that_does_not_exist.xlsx")


def _make_elevations():
    from app.converters.team_parser import TEAMElevation

    return [
        TEAMElevation(label="A", left=["100"], cntr=["101"], rght=["102"], has_data=True),
        TEAMElevation(label="B", left=[""], cntr=[""], rght=[""], has_data=False),
        TEAMElevation(label="C", left=["200"], cntr=["201"], rght=["202"], has_data=True),
        TEAMElevation(label="D", left=[""], cntr=[""], rght=[""], has_data=False),
    ]


def test_filter_elevations_by_data_excludes_blank_by_default():
    from app.converters.team_parser import filter_elevations_by_data

    elevs = _make_elevations()
    result = filter_elevations_by_data(elevs, include_blank=False)
    assert [e.label for e in result] == ["A", "C"]
    assert all(e.has_data for e in result)


def test_filter_elevations_by_data_includes_blank_when_requested():
    from app.converters.team_parser import filter_elevations_by_data

    elevs = _make_elevations()
    result = filter_elevations_by_data(elevs, include_blank=True)
    assert result == elevs
