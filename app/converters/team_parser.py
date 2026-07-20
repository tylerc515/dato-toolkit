"""TEAM xlsx inspection file parser."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


class TEAMParseError(Exception):
    """Raised when a TEAM xlsx file cannot be parsed."""


@dataclass
class TEAMElevation:
    label: str
    left: list[str]
    cntr: list[str]
    rght: list[str]
    has_data: bool   # True if any reading cell in this block is non-blank


@dataclass
class TEAMParseResult:
    tube_numbers: list[int]
    numbering_direction: str   # "Left-to-Right" or "Right-to-Left"
    num_tubes: int
    elevations: list[TEAMElevation]
    comment_codes_found: set[str]      # every unique non-numeric symbol seen in readings


@dataclass
class TEAMConversionInput:
    """A TEAMParseResult plus batch metadata and the per-file section
    name, shaped exactly like what write_standard_format() reads off an
    ATSParseResult. Built fresh per file at conversion time."""
    company_name: str
    mill_location: str
    boiler_name: str
    inspection_date: str      # "Month YYYY", already formatted
    boiler_section: str       # per-file, from the editable section field
    nde_laboratory: str
    num_tubes: int
    numbering_direction: str
    tube_numbers: list[int]
    elevations: list[TEAMElevation]


_SUFFIX_PATTERN = re.compile(r"^(\d+)([A-Za-z])$")


def convert_team_reading(value: object) -> str:
    """Convert a TEAM cell value to a Standard Format reading string.

    int/float (readings are already whole-number thousandths) -> zero-padded
        to 3 digits, written through at natural width for out-of-range
        values (e.g. 2791 -> "2791").
    None or blank -> ""
    single-character string (comment code symbol) -> unchanged
    digits + trailing letter suffix (e.g. "14V") -> zero-padded digits with
        the suffix preserved ("014V")
    anything else -> stripped and returned unchanged
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return f"{int(value):03d}"
    stripped = str(value).strip()
    if not stripped:
        return ""
    if len(stripped) == 1:
        return stripped
    match = _SUFFIX_PATTERN.match(stripped)
    if match:
        digits, suffix = match.groups()
        return f"{int(digits):03d}{suffix}"
    return stripped


def _is_numeric_reading(stripped: str) -> bool:
    """True if a stripped string reading is numeric (plain or suffix-lettered)."""
    if _SUFFIX_PATTERN.match(stripped):
        return True
    try:
        float(stripped)
        return True
    except ValueError:
        return False


def parse_team_file(filepath: str | Path) -> TEAMParseResult:
    """Parse a TEAM xlsx inspection file into a TEAMParseResult.

    Raises:
        TEAMParseError: if the file cannot be opened or has no tube numbers.
    """
    path = Path(filepath)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise TEAMParseError(f"Cannot open '{path.name}': {exc}") from exc

    try:
        ws = wb.active

        # Row 1: tube numbers starting at column C (column 3)
        tube_numbers: list[int] = []
        col = 3
        while True:
            val = ws.cell(row=1, column=col).value
            if isinstance(val, bool) or not isinstance(val, (int, float)):
                break
            tube_numbers.append(int(val))
            col += 1

        num_tubes = len(tube_numbers)
        if num_tubes == 0:
            raise TEAMParseError(f"'{path.name}': no tube numbers found in row 1")

        numbering_direction = (
            "Left-to-Right" if tube_numbers[0] == 1 else "Right-to-Left"
        )

        last_reading_col = 2 + num_tubes  # readings run from column C to here

        elevations: list[TEAMElevation] = []
        comment_codes_found: set[str] = set()

        row_idx = 2
        while row_idx + 2 <= ws.max_row:
            block_rows = (row_idx, row_idx + 1, row_idx + 2)

            label_parts = []
            for r in block_rows:
                part = ws.cell(row=r, column=1).value
                if part is not None:
                    stripped_part = str(part).strip()
                    if stripped_part:
                        label_parts.append(stripped_part)
            label = " ".join(label_parts)

            raw_rows = [
                [ws.cell(row=r, column=c).value for c in range(3, last_reading_col + 1)]
                for r in block_rows
            ]

            all_blank = all(
                cell is None or (isinstance(cell, str) and not cell.strip())
                for row in raw_rows
                for cell in row
            )
            has_data = not all_blank

            left = [convert_team_reading(v) for v in raw_rows[0]]
            cntr = [convert_team_reading(v) for v in raw_rows[1]]
            rght = [convert_team_reading(v) for v in raw_rows[2]]

            for row in raw_rows:
                for cell in row:
                    if not isinstance(cell, str):
                        continue
                    cell_stripped = cell.strip()
                    if cell_stripped and not _is_numeric_reading(cell_stripped):
                        comment_codes_found.add(cell_stripped)

            elevations.append(TEAMElevation(
                label=label,
                left=left,
                cntr=cntr,
                rght=rght,
                has_data=has_data,
            ))
            row_idx += 3

        return TEAMParseResult(
            tube_numbers=tube_numbers,
            numbering_direction=numbering_direction,
            num_tubes=num_tubes,
            elevations=elevations,
            comment_codes_found=comment_codes_found,
        )
    finally:
        wb.close()


def filter_elevations_by_data(
    elevations: list[TEAMElevation],
    include_blank: bool,
) -> list[TEAMElevation]:
    """If include_blank is True, return all elevations unchanged.
    If False, return only elevations whose has_data is True."""
    if include_blank:
        return list(elevations)
    return [e for e in elevations if e.has_data]
